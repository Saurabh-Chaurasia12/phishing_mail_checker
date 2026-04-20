"""Parse Neon offline exports into lightweight Python data structures.

The parser is intentionally tolerant of column naming variations because Neon
CSV exports may include unit suffixes such as ``[ns]`` or ``[px]``.
"""

from __future__ import annotations

import csv
import json
import os
import re
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional


def _normalize_key(key: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", key.strip().lower()).strip("_")


def _normalize_row(row: Dict[str, str]) -> Dict[str, str]:
    return {_normalize_key(k): (v or "").strip() for k, v in row.items()}


def _as_float(value: Optional[str], default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _as_int(value: Optional[str], default: int = 0) -> int:
    if value is None or value == "":
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _read_csv_rows(path: str) -> List[Dict[str, str]]:
    if not os.path.isfile(path):
        return []
    with open(path, "r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return [_normalize_row(row) for row in reader]


def _read_xlsx_rows(path: str) -> List[Dict[str, str]]:
    if not os.path.isfile(path):
        return []
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to read xlsx Neon fallback data. Install with: pip install openpyxl"
        ) from exc

    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []

    header_keys = [_normalize_key(str(h)) if h is not None else "" for h in header]
    records: List[Dict[str, str]] = []
    for row in rows_iter:
        rec: Dict[str, str] = {}
        for i, value in enumerate(row):
            key = header_keys[i] if i < len(header_keys) and header_keys[i] else f"col_{i}"
            rec[key] = "" if value is None else str(value)
        records.append(rec)
    return records


def _pick_key(row: Dict[str, str], candidates: List[str]) -> Optional[str]:
    keys = set(row.keys())
    for cand in candidates:
        if cand in keys:
            return cand
    return None


def _parse_gaze_xlsx(path: str) -> List[NeonGazeSample]:
    rows = _read_xlsx_rows(path)
    if not rows:
        return []

    samples: List[NeonGazeSample] = []
    for row in rows:
        ts_key = _pick_key(row, ["timestamp_ns", "timestamp", "time_ns", "time", "t"])
        x_key = _pick_key(row, ["gaze_x_px", "gaze_x", "x", "screen_x", "coord_x"])
        y_key = _pick_key(row, ["gaze_y_px", "gaze_y", "y", "screen_y", "coord_y"])
        conf_key = _pick_key(row, ["confidence", "conf", "score"])

        if ts_key is None or x_key is None or y_key is None:
            ordered_keys = list(row.keys())
            if len(ordered_keys) >= 3:
                ts_key = ordered_keys[0]
                x_key = ordered_keys[1]
                y_key = ordered_keys[2]
            else:
                continue

        timestamp_raw = _as_float(row.get(ts_key), 0.0)
        # Handle likely seconds timestamps by converting to ns.
        timestamp_ns = int(timestamp_raw if timestamp_raw > 1e12 else timestamp_raw * 1_000_000_000)
        samples.append(
            NeonGazeSample(
                timestamp_ns=timestamp_ns,
                gaze_x_px=_as_float(row.get(x_key), 0.0),
                gaze_y_px=_as_float(row.get(y_key), 0.0),
                confidence=_as_float(row.get(conf_key), 1.0) if conf_key else 1.0,
                raw=row,
            )
        )

    return samples


@dataclass
class NeonGazeSample:
    timestamp_ns: int
    gaze_x_px: float
    gaze_y_px: float
    confidence: float
    worn: float = 1.0
    fixation_id: Optional[int] = None
    blink_id: Optional[int] = None
    raw: Dict[str, str] = field(default_factory=dict)


@dataclass
class NeonFixationSample:
    fixation_id: int
    start_timestamp_ns: int
    end_timestamp_ns: int
    duration_ms: float
    fixation_x_px: float
    fixation_y_px: float
    raw: Dict[str, str] = field(default_factory=dict)


@dataclass
class NeonIMUSample:
    timestamp_ns: int
    pitch_deg: Optional[float] = None
    yaw_deg: Optional[float] = None
    roll_deg: Optional[float] = None
    raw: Dict[str, str] = field(default_factory=dict)


@dataclass
class NeonBlinkSample:
    blink_id: int
    start_timestamp_ns: int
    end_timestamp_ns: int
    duration_ms: float
    raw: Dict[str, str] = field(default_factory=dict)


@dataclass
class NeonRecording:
    recording_dir: str
    info: Dict[str, Any] = field(default_factory=dict)
    gaze_samples: List[NeonGazeSample] = field(default_factory=list)
    fixations: List[NeonFixationSample] = field(default_factory=list)
    imu_samples: List[NeonIMUSample] = field(default_factory=list)
    blinks: List[NeonBlinkSample] = field(default_factory=list)
    eye_state_rows: List[Dict[str, str]] = field(default_factory=list)

    @property
    def start_time_ns(self) -> int:
        if self.info.get("start_time") is not None:
            return _as_int(str(self.info["start_time"]), 0)

        timestamps: List[int] = []
        for sample in self.gaze_samples:
            timestamps.append(sample.timestamp_ns)
        for sample in self.fixations:
            timestamps.append(sample.start_timestamp_ns)
        for sample in self.imu_samples:
            timestamps.append(sample.timestamp_ns)
        for sample in self.blinks:
            timestamps.append(sample.start_timestamp_ns)
        return min(timestamps) if timestamps else 0

    @property
    def end_time_ns(self) -> int:
        timestamps: List[int] = []
        for sample in self.gaze_samples:
            timestamps.append(sample.timestamp_ns)
        for sample in self.fixations:
            timestamps.append(sample.end_timestamp_ns)
        for sample in self.imu_samples:
            timestamps.append(sample.timestamp_ns)
        for sample in self.blinks:
            timestamps.append(sample.end_timestamp_ns)
        return max(timestamps) if timestamps else self.start_time_ns


def _load_info_json(recording_dir: str) -> Dict[str, Any]:
    path = os.path.join(recording_dir, "info.json")
    if not os.path.isfile(path):
        return {}
    with open(path, "r", encoding="utf-8") as handle:
        return json.load(handle)


def _parse_gaze_csv(recording_dir: str) -> List[NeonGazeSample]:
    rows = _read_csv_rows(os.path.join(recording_dir, "gaze.csv"))
    if not rows:
        xlsx_path = os.path.join(recording_dir, "time.xlsx")
        xlsx_samples = _parse_gaze_xlsx(xlsx_path)
        if xlsx_samples:
            return xlsx_samples

    samples: List[NeonGazeSample] = []
    for row in rows:
        timestamp_ns = _as_int(row.get("timestamp_ns") or row.get("timestamp"))
        gaze_x_px = _as_float(row.get("gaze_x_px") or row.get("gaze_x"))
        gaze_y_px = _as_float(row.get("gaze_y_px") or row.get("gaze_y"))
        confidence = _as_float(row.get("confidence"), 0.0)
        worn = _as_float(row.get("worn"), 1.0)
        fixation_id = row.get("fixation_id")
        blink_id = row.get("blink_id")
        samples.append(
            NeonGazeSample(
                timestamp_ns=timestamp_ns,
                gaze_x_px=gaze_x_px,
                gaze_y_px=gaze_y_px,
                confidence=confidence,
                worn=worn,
                fixation_id=_as_int(fixation_id, default=-1) if fixation_id not in (None, "") else None,
                blink_id=_as_int(blink_id, default=-1) if blink_id not in (None, "") else None,
                raw=row,
            )
        )
    return samples


def _parse_fixations_csv(recording_dir: str) -> List[NeonFixationSample]:
    rows = _read_csv_rows(os.path.join(recording_dir, "fixations.csv"))
    samples: List[NeonFixationSample] = []
    for row in rows:
        samples.append(
            NeonFixationSample(
                fixation_id=_as_int(row.get("fixation_id")),
                start_timestamp_ns=_as_int(row.get("start_timestamp_ns") or row.get("start_timestamp")),
                end_timestamp_ns=_as_int(row.get("end_timestamp_ns") or row.get("end_timestamp")),
                duration_ms=_as_float(row.get("duration_ms") or row.get("duration")),
                fixation_x_px=_as_float(row.get("fixation_x_px") or row.get("fixation_x")),
                fixation_y_px=_as_float(row.get("fixation_y_px") or row.get("fixation_y")),
                raw=row,
            )
        )
    return samples


def _parse_imu_csv(recording_dir: str) -> List[NeonIMUSample]:
    rows = _read_csv_rows(os.path.join(recording_dir, "imu.csv"))
    samples: List[NeonIMUSample] = []
    for row in rows:
        samples.append(
            NeonIMUSample(
                timestamp_ns=_as_int(row.get("timestamp_ns") or row.get("timestamp")),
                pitch_deg=_as_float(row.get("pitch_deg") or row.get("pitch"), default=None),
                yaw_deg=_as_float(row.get("yaw_deg") or row.get("yaw"), default=None),
                roll_deg=_as_float(row.get("roll_deg") or row.get("roll"), default=None),
                raw=row,
            )
        )
    return samples


def _parse_blinks_csv(recording_dir: str) -> List[NeonBlinkSample]:
    rows = _read_csv_rows(os.path.join(recording_dir, "blinks.csv"))
    samples: List[NeonBlinkSample] = []
    for row in rows:
        samples.append(
            NeonBlinkSample(
                blink_id=_as_int(row.get("blink_id")),
                start_timestamp_ns=_as_int(row.get("start_timestamp_ns") or row.get("start_timestamp")),
                end_timestamp_ns=_as_int(row.get("end_timestamp_ns") or row.get("end_timestamp")),
                duration_ms=_as_float(row.get("duration_ms") or row.get("duration")),
                raw=row,
            )
        )
    return samples


def _parse_eye_state_rows(recording_dir: str) -> List[Dict[str, str]]:
    return _read_csv_rows(os.path.join(recording_dir, "3d_eye_states.csv"))


def parse_recording(recording_dir: str) -> NeonRecording:
    """Load the Neon recording export that lives in ``recording_dir``."""
    return NeonRecording(
        recording_dir=recording_dir,
        info=_load_info_json(recording_dir),
        gaze_samples=_parse_gaze_csv(recording_dir),
        fixations=_parse_fixations_csv(recording_dir),
        imu_samples=_parse_imu_csv(recording_dir),
        blinks=_parse_blinks_csv(recording_dir),
        eye_state_rows=_parse_eye_state_rows(recording_dir),
    )
