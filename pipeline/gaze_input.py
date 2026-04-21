from __future__ import annotations

from dataclasses import dataclass
from typing import Iterable, Optional


@dataclass
class GazeSample:
    timestamp_s: float
    x: float
    y: float
    confidence: float = 1.0


def _normalize_key(value: object) -> str:
    text = "" if value is None else str(value)
    return "".join(ch.lower() if ch.isalnum() else "_" for ch in text).strip("_")


def _pick_key(keys: Iterable[str], candidates: list[str]) -> Optional[str]:
    key_set = set(keys)
    for candidate in candidates:
        if candidate in key_set:
            return candidate
    return None


def _as_float(value: object, default: float = 0.0) -> float:
    if value in (None, ""):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _to_seconds(raw_timestamp: float) -> float:
    if raw_timestamp > 1e14:
        return raw_timestamp / 1_000_000_000.0
    if raw_timestamp > 1e9:
        return raw_timestamp / 1000.0
    return raw_timestamp


def load_gaze_samples_from_xlsx(path: str) -> list[GazeSample]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read gaze xlsx files.") from exc

    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return []

    normalized_header = [_normalize_key(cell) for cell in header]
    ts_key = _pick_key(normalized_header, ["time_stamp", "timestamp", "time", "t"])
    x_key = _pick_key(normalized_header, ["x", "gaze_x", "screen_x", "coord_x"])
    y_key = _pick_key(normalized_header, ["y", "gaze_y", "screen_y", "coord_y"])
    conf_key = _pick_key(normalized_header, ["confidence", "conf", "score"])

    if ts_key is None or x_key is None or y_key is None:
        raise ValueError("time.xlsx must contain timestamp, x, and y columns.")

    ts_idx = normalized_header.index(ts_key)
    x_idx = normalized_header.index(x_key)
    y_idx = normalized_header.index(y_key)
    conf_idx = normalized_header.index(conf_key) if conf_key is not None else None

    samples: list[GazeSample] = []
    for row in rows:
        raw_ts = _as_float(row[ts_idx] if ts_idx < len(row) else None, 0.0)
        x = _as_float(row[x_idx] if x_idx < len(row) else None, 0.0)
        y = _as_float(row[y_idx] if y_idx < len(row) else None, 0.0)
        conf = _as_float(row[conf_idx] if conf_idx is not None and conf_idx < len(row) else None, 1.0)
        samples.append(
            GazeSample(
                timestamp_s=_to_seconds(raw_ts),
                x=x,
                y=y,
                confidence=conf,
            )
        )

    return sorted(samples, key=lambda sample: sample.timestamp_s)

